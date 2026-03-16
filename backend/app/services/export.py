"""Export service for deals, properties, and portfolio reports.

Produces CSV, XLSX, and HTML (print-to-PDF) exports of saved deals with
all financial metrics.
"""

from __future__ import annotations

import csv
import io
import logging
from datetime import datetime, timezone
from typing import Any
from uuid import UUID

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.property import Property
from app.models.saved_deal import SavedDeal
from app.models.user import User

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

DEAL_COLUMNS: list[tuple[str, str]] = [
    ("address", "Address"),
    ("city", "City"),
    ("state", "State"),
    ("zip_code", "Zip Code"),
    ("property_type", "Property Type"),
    ("price", "List Price"),
    ("arv", "After Repair Value"),
    ("rehab_cost", "Rehab Cost"),
    ("monthly_rent", "Monthly Rent"),
    ("annual_rent", "Annual Rent"),
    ("cap_rate", "Cap Rate (%)"),
    ("cash_flow", "Monthly Cash Flow"),
    ("coc_return", "Cash-on-Cash Return (%)"),
    ("dscr", "DSCR"),
    ("brrrr_score", "BRRRR Score"),
    ("investment_score", "Investment Score"),
    ("notes", "Notes"),
    ("is_favorite", "Favorite"),
    ("saved_at", "Saved Date"),
]


class ExportService:
    """Generate CSV, XLSX, and HTML exports of saved deals."""

    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    # ------------------------------------------------------------------
    # Data loading
    # ------------------------------------------------------------------

    async def _load_deals(
        self, user_id: UUID, deal_ids: list[UUID] | None = None
    ) -> list[dict[str, Any]]:
        """Load saved deals with property data and flatten into dicts."""
        stmt = (
            select(SavedDeal)
            .options(selectinload(SavedDeal.property))
            .where(SavedDeal.user_id == user_id)
            .order_by(SavedDeal.created_at.desc())
        )
        if deal_ids:
            stmt = stmt.where(SavedDeal.id.in_(deal_ids))

        result = await self.db.execute(stmt)
        saved_deals = result.scalars().all()

        rows: list[dict[str, Any]] = []
        for sd in saved_deals:
            prop: Property | None = sd.property
            if not prop:
                continue

            price = float(sd.custom_arv or prop.purchase_price or 0)
            arv = float(sd.custom_arv or prop.current_value or 0)
            rehab = float(sd.custom_rehab or 0)
            rent = float(sd.custom_rent or 0)
            annual_rent = rent * 12

            # Financial calculations
            cap_rate = (annual_rent / price * 100) if price > 0 else 0
            monthly_expenses = (
                float(prop.mortgage_payment or 0)
                + float(prop.insurance_cost or 0)
                + float(prop.tax_annual or 0) / 12
            )
            cash_flow = rent - monthly_expenses
            total_investment = price + rehab
            coc_return = (
                (cash_flow * 12 / total_investment * 100) if total_investment > 0 else 0
            )
            dscr = (
                rent / float(prop.mortgage_payment or 1) if prop.mortgage_payment else 0
            )
            brrrr_score = _compute_brrrr_score(price, arv, rehab, rent)
            investment_score = _compute_investment_score(
                cap_rate, cash_flow, coc_return, dscr
            )

            rows.append(
                {
                    "address": f"{prop.address_line1}"
                    + (f" {prop.address_line2}" if prop.address_line2 else ""),
                    "city": prop.city,
                    "state": prop.state,
                    "zip_code": prop.zip_code,
                    "property_type": (
                        prop.property_type.value if prop.property_type else ""
                    ),
                    "price": round(price, 2),
                    "arv": round(arv, 2),
                    "rehab_cost": round(rehab, 2),
                    "monthly_rent": round(rent, 2),
                    "annual_rent": round(annual_rent, 2),
                    "cap_rate": round(cap_rate, 2),
                    "cash_flow": round(cash_flow, 2),
                    "coc_return": round(coc_return, 2),
                    "dscr": round(dscr, 2),
                    "brrrr_score": round(brrrr_score, 1),
                    "investment_score": investment_score,
                    "notes": sd.notes or "",
                    "is_favorite": "Yes" if sd.is_favorite else "No",
                    "saved_at": sd.created_at.strftime("%Y-%m-%d")
                    if sd.created_at
                    else "",
                }
            )

        return rows

    # ------------------------------------------------------------------
    # CSV export
    # ------------------------------------------------------------------

    async def export_deals_csv(
        self, user_id: UUID, deal_ids: list[UUID] | None = None
    ) -> bytes:
        """Export saved deals as a UTF-8 CSV file.

        Returns raw bytes suitable for streaming as a file download.
        """
        rows = await self._load_deals(user_id, deal_ids)

        buf = io.StringIO()
        writer = csv.writer(buf)

        # Header row
        writer.writerow([col_label for _, col_label in DEAL_COLUMNS])

        # Data rows
        for row in rows:
            writer.writerow([row.get(col_key, "") for col_key, _ in DEAL_COLUMNS])

        csv_bytes = buf.getvalue().encode("utf-8-sig")  # BOM for Excel compat
        logger.info(
            "CSV export: user=%s deals=%d size=%d bytes",
            user_id,
            len(rows),
            len(csv_bytes),
        )
        return csv_bytes

    # ------------------------------------------------------------------
    # XLSX export
    # ------------------------------------------------------------------

    async def export_deals_xlsx(
        self, user_id: UUID, deal_ids: list[UUID] | None = None
    ) -> bytes:
        """Export saved deals as an Excel workbook with formatted sheets.

        Sheets:
        1. Summary — overview metrics
        2. Financial Analysis — all deals with financial columns
        3. Market Data — deals grouped by market
        """
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        rows = await self._load_deals(user_id, deal_ids)
        wb = openpyxl.Workbook()

        # ---- Sheet 1: Summary ----
        ws_summary = wb.active
        ws_summary.title = "Summary"

        summary_header_font = Font(bold=True, size=14)
        ws_summary["A1"] = "RealDeal AI — Deal Export"
        ws_summary["A1"].font = summary_header_font
        ws_summary["A2"] = (
            f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
        )
        ws_summary["A3"] = f"Total Deals: {len(rows)}"

        if rows:
            prices = [r["price"] for r in rows if r["price"] > 0]
            cap_rates = [r["cap_rate"] for r in rows if r["cap_rate"] > 0]
            cash_flows = [r["cash_flow"] for r in rows]

            summary_data = [
                ("Metric", "Value"),
                ("Total Deals", len(rows)),
                (
                    "Average Price",
                    f"${sum(prices) / len(prices):,.0f}" if prices else "N/A",
                ),
                (
                    "Average Cap Rate",
                    f"{sum(cap_rates) / len(cap_rates):.2f}%" if cap_rates else "N/A",
                ),
                (
                    "Average Cash Flow",
                    f"${sum(cash_flows) / len(cash_flows):,.0f}"
                    if cash_flows
                    else "N/A",
                ),
                ("Favorites", sum(1 for r in rows if r["is_favorite"] == "Yes")),
            ]

            for r_idx, (label, value) in enumerate(summary_data, start=5):
                ws_summary.cell(row=r_idx, column=1, value=label).font = Font(bold=True)
                ws_summary.cell(row=r_idx, column=2, value=value)

        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 25

        # ---- Sheet 2: Financial Analysis ----
        ws_finance = wb.create_sheet("Financial Analysis")

        header_fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col_idx, (_, col_label) in enumerate(DEAL_COLUMNS, start=1):
            cell = ws_finance.cell(row=1, column=col_idx, value=col_label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        money_format = "#,##0.00"
        pct_format = '0.00"%"'

        money_cols = {
            "price",
            "arv",
            "rehab_cost",
            "monthly_rent",
            "annual_rent",
            "cash_flow",
        }
        pct_cols = {"cap_rate", "coc_return"}

        for r_idx, row in enumerate(rows, start=2):
            for col_idx, (col_key, _) in enumerate(DEAL_COLUMNS, start=1):
                value = row.get(col_key, "")
                cell = ws_finance.cell(row=r_idx, column=col_idx, value=value)

                if col_key in money_cols and isinstance(value, (int, float)):
                    cell.number_format = money_format
                elif col_key in pct_cols and isinstance(value, (int, float)):
                    cell.number_format = pct_format

        # Auto-fit column widths (approximate)
        for col_idx, (col_key, col_label) in enumerate(DEAL_COLUMNS, start=1):
            max_len = len(col_label)
            for row in rows[:50]:
                val_len = len(str(row.get(col_key, "")))
                if val_len > max_len:
                    max_len = val_len
            ws_finance.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = min(max_len + 4, 40)

        # ---- Sheet 3: Market Data ----
        ws_market = wb.create_sheet("Market Data")

        markets: dict[str, list[dict[str, Any]]] = {}
        for row in rows:
            market_key = f"{row['city']}, {row['state']}"
            markets.setdefault(market_key, []).append(row)

        current_row = 1
        for market_name, market_rows in sorted(markets.items()):
            ws_market.cell(row=current_row, column=1, value=market_name).font = Font(
                bold=True, size=12
            )
            current_row += 1

            market_prices = [r["price"] for r in market_rows if r["price"] > 0]
            market_caps = [r["cap_rate"] for r in market_rows if r["cap_rate"] > 0]

            stats = [
                ("Deals", len(market_rows)),
                (
                    "Avg Price",
                    f"${sum(market_prices) / len(market_prices):,.0f}"
                    if market_prices
                    else "N/A",
                ),
                (
                    "Avg Cap Rate",
                    f"{sum(market_caps) / len(market_caps):.2f}%"
                    if market_caps
                    else "N/A",
                ),
            ]
            for label, value in stats:
                ws_market.cell(row=current_row, column=1, value=label).font = Font(
                    bold=True
                )
                ws_market.cell(row=current_row, column=2, value=value)
                current_row += 1

            current_row += 1

        ws_market.column_dimensions["A"].width = 25
        ws_market.column_dimensions["B"].width = 25

        # Write to bytes
        output = io.BytesIO()
        wb.save(output)
        xlsx_bytes = output.getvalue()

        logger.info(
            "XLSX export: user=%s deals=%d size=%d bytes",
            user_id,
            len(rows),
            len(xlsx_bytes),
        )
        return xlsx_bytes

    # ------------------------------------------------------------------
    # Portfolio report (HTML for print-to-PDF)
    # ------------------------------------------------------------------

    async def export_portfolio_report(self, user_id: UUID) -> bytes:
        """Generate an HTML portfolio summary that can be printed to PDF.

        Returns UTF-8 encoded HTML bytes.
        """
        rows = await self._load_deals(user_id)

        # Load user info
        result = await self.db.execute(select(User).where(User.id == user_id))
        user = result.scalar_one_or_none()
        user_name = user.full_name if user else "Investor"

        # Aggregate stats
        total_value = sum(r["price"] for r in rows)
        total_rent = sum(r["monthly_rent"] for r in rows)
        avg_cap_rate = sum(r["cap_rate"] for r in rows) / len(rows) if rows else 0
        avg_score = sum(r["investment_score"] for r in rows) / len(rows) if rows else 0
        total_cash_flow = sum(r["cash_flow"] for r in rows)
        sum(1 for r in rows if r["is_favorite"] == "Yes")

        # Group by market
        markets: dict[str, list[dict[str, Any]]] = {}
        for row in rows:
            key = f"{row['city']}, {row['state']}"
            markets.setdefault(key, []).append(row)

        now = datetime.now(timezone.utc).strftime("%B %d, %Y")

        # Build deals table rows
        deal_rows_html = ""
        for i, row in enumerate(rows, 1):
            score_color = (
                "#27ae60"
                if row["investment_score"] >= 70
                else "#f39c12"
                if row["investment_score"] >= 40
                else "#e74c3c"
            )
            deal_rows_html += f"""
            <tr>
                <td>{i}</td>
                <td>{row["address"]}<br><small>{row["city"]}, {row["state"]} {row["zip_code"]}</small></td>
                <td>${row["price"]:,.0f}</td>
                <td>${row["arv"]:,.0f}</td>
                <td>${row["monthly_rent"]:,.0f}</td>
                <td>{row["cap_rate"]:.1f}%</td>
                <td>${row["cash_flow"]:,.0f}</td>
                <td>{row["coc_return"]:.1f}%</td>
                <td>{row["dscr"]:.2f}</td>
                <td style="color:{score_color};font-weight:bold">{row["investment_score"]}</td>
            </tr>"""

        # Market summary rows
        market_rows_html = ""
        for market_name, market_rows in sorted(
            markets.items(), key=lambda x: len(x[1]), reverse=True
        ):
            m_prices = [r["price"] for r in market_rows if r["price"] > 0]
            m_caps = [r["cap_rate"] for r in market_rows if r["cap_rate"] > 0]
            market_rows_html += f"""
            <tr>
                <td>{market_name}</td>
                <td>{len(market_rows)}</td>
                <td>${sum(m_prices) / len(m_prices):,.0f}</td>
                <td>{sum(m_caps) / len(m_caps):.1f}%</td>
            </tr>"""

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Portfolio Report — {user_name}</title>
<style>
    @page {{ margin: 0.75in; }}
    body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; color: #2c3e50; margin: 0; padding: 20px; font-size: 12px; }}
    h1 {{ color: #1a5276; margin-bottom: 4px; }}
    h2 {{ color: #1f4e79; border-bottom: 2px solid #1f4e79; padding-bottom: 4px; margin-top: 30px; }}
    .subtitle {{ color: #7f8c8d; font-size: 14px; margin-bottom: 20px; }}
    .stats-grid {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 15px; margin: 20px 0; }}
    .stat-card {{ background: #f8f9fa; border-radius: 8px; padding: 15px; text-align: center; border: 1px solid #e9ecef; }}
    .stat-value {{ font-size: 22px; font-weight: bold; color: #1a5276; }}
    .stat-label {{ font-size: 11px; color: #7f8c8d; margin-top: 4px; }}
    table {{ width: 100%; border-collapse: collapse; margin: 15px 0; font-size: 11px; }}
    th {{ background: #1f4e79; color: white; padding: 8px 6px; text-align: left; font-size: 10px; }}
    td {{ padding: 6px; border-bottom: 1px solid #e9ecef; }}
    tr:nth-child(even) {{ background: #f8f9fa; }}
    .footer {{ margin-top: 30px; text-align: center; color: #95a5a6; font-size: 10px; border-top: 1px solid #e9ecef; padding-top: 10px; }}
</style>
</head>
<body>
    <h1>RealDeal AI Portfolio Report</h1>
    <p class="subtitle">Prepared for {user_name} on {now}</p>

    <h2>Portfolio Overview</h2>
    <div class="stats-grid">
        <div class="stat-card">
            <div class="stat-value">{len(rows)}</div>
            <div class="stat-label">Saved Deals</div>
        </div>
        <div class="stat-card">
            <div class="stat-value">${total_value:,.0f}</div>
            <div class="stat-label">Total Portfolio Value</div>
        </div>
        <div class="stat-card">
            <div class="stat-value">${total_rent:,.0f}/mo</div>
            <div class="stat-label">Total Monthly Rent</div>
        </div>
        <div class="stat-card">
            <div class="stat-value">{avg_cap_rate:.1f}%</div>
            <div class="stat-label">Average Cap Rate</div>
        </div>
        <div class="stat-card">
            <div class="stat-value">${total_cash_flow:,.0f}/mo</div>
            <div class="stat-label">Total Monthly Cash Flow</div>
        </div>
        <div class="stat-card">
            <div class="stat-value">{avg_score:.0f}</div>
            <div class="stat-label">Average Investment Score</div>
        </div>
    </div>

    <h2>Deal Details</h2>
    <table>
        <thead>
            <tr>
                <th>#</th>
                <th>Property</th>
                <th>Price</th>
                <th>ARV</th>
                <th>Rent/Mo</th>
                <th>Cap Rate</th>
                <th>Cash Flow</th>
                <th>CoC Return</th>
                <th>DSCR</th>
                <th>Score</th>
            </tr>
        </thead>
        <tbody>{deal_rows_html}
        </tbody>
    </table>

    <h2>Market Breakdown</h2>
    <table>
        <thead>
            <tr>
                <th>Market</th>
                <th>Deals</th>
                <th>Avg Price</th>
                <th>Avg Cap Rate</th>
            </tr>
        </thead>
        <tbody>{market_rows_html}
        </tbody>
    </table>

    <div class="footer">
        Generated by RealDeal AI &mdash; {now}<br>
        This report is for informational purposes only and does not constitute financial advice.
    </div>
</body>
</html>"""

        html_bytes = html.encode("utf-8")
        logger.info(
            "Portfolio report: user=%s deals=%d size=%d bytes",
            user_id,
            len(rows),
            len(html_bytes),
        )
        return html_bytes


# ---------------------------------------------------------------------------
# Scoring helpers
# ---------------------------------------------------------------------------


def _compute_brrrr_score(
    price: float, arv: float, rehab: float, monthly_rent: float
) -> float:
    """Compute a 0-100 BRRRR strategy score.

    Factors: ARV-to-cost ratio, rent-to-price ratio, equity capture potential.
    """
    total_cost = price + rehab
    if total_cost <= 0:
        return 0.0

    # 70% rule: total cost should be <= 70% of ARV
    cost_ratio = total_cost / arv if arv > 0 else 1.0
    rule_70_score = max(0, min(40, (0.7 - cost_ratio) / 0.7 * 40 + 20))

    # Rent-to-price (1% rule: monthly rent >= 1% of price)
    rent_ratio = (monthly_rent / price * 100) if price > 0 else 0
    rent_score = max(0, min(30, rent_ratio / 1.0 * 30))

    # Equity capture: (ARV - total_cost) / total_cost
    equity_pct = ((arv - total_cost) / total_cost * 100) if total_cost > 0 else 0
    equity_score = max(0, min(30, equity_pct / 30 * 30))

    return round(rule_70_score + rent_score + equity_score, 1)


def _compute_investment_score(
    cap_rate: float, cash_flow: float, coc_return: float, dscr: float
) -> int:
    """Compute a 0-100 overall investment score."""
    # Cap rate: 10+ is excellent (30 pts)
    cap_score = min(30, cap_rate / 10 * 30)

    # Cash flow: $500+/mo is excellent (25 pts)
    cf_score = min(25, max(0, cash_flow / 500 * 25))

    # CoC return: 12%+ is excellent (25 pts)
    coc_score = min(25, coc_return / 12 * 25)

    # DSCR: 1.5+ is excellent (20 pts)
    dscr_score = min(20, dscr / 1.5 * 20) if dscr > 0 else 0

    return min(100, max(0, int(cap_score + cf_score + coc_score + dscr_score)))
